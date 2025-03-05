import os
import openpyxl
from openpyxl import load_workbook
import numpy as np


directories = ["AllPassBiQuadFilter_aucs", "BandBiQuadFilter_aucs", "BandPassBiQuadFilter_aucs", 
               "BandRejectBiQuadFilter_aucs", "HighPassBiQuadFilter_aucs", "LowPassBiQuadFilter_aucs",
               "TrebleBiQuadFilter_aucs"] 
directories_avg_mel = ["AllPassBiQuadFilter_Avg_Mel_aucs", "BandBiQuadFilter_Avg_Mel_aucs", "BandPassBiQuadFilter_Avg_Mel_aucs", 
               "BandRejectBiQuadFilter_Avg_Mel_aucs", "HighPassBiQuadFilter_Avg_Mel_aucs", "LowPassBiQuadFilter_Avg_Mel_aucs",
               "TrebleBiQuadFilter_Avg_Mel_aucs"]

directories_butter = ["LFilterFilter_4_highpass_Avg_Mel_aucs", "LFilterFilter_4_lowpass_Avg_Mel_aucs", 
                      "LFilterFilter_4_bandpass_Avg_Mel_aucs", "LFilterFilter_4_bandreject_Avg_Mel_aucs",
                      "FiltFiltFilter_4_highpass_Avg_Mel_aucs", "FiltFiltFilter_4_lowpass_Avg_Mel_aucs",
                      "FiltFiltFilter_4_bandpass_Avg_Mel_aucs", "FiltFiltFilter_4_bandreject_Avg_Mel_aucs"]

directories = directories_avg_mel#butter
directories = ["BandPassBiQuadFilter_Avg_MFCC_aucs", "BandPassBiQuadFilter_Avg_Spec_aucs", "BandPassBiQuadFilter_Avg_Mel_aucs"]


# Function to process each Excel file
def process_excel_file(file_path):
    workbook = load_workbook(file_path, data_only=True)
    sheet_names = workbook.sheetnames

    all_values = []
    
    for sheet_name in sheet_names:
        if sheet_name.startswith('LJSpeech'):
            #print("Attributing original real data is skipped") 
            continue 
        sheet = workbook[sheet_name]
        # Read values from B2 to B7
        values = [sheet[f'B{i}'].value for i in range(2, 8)]
        all_values.extend(values)

    # Calculate average and minimum
    average_value = np.mean(all_values)
    min_value = np.min(all_values)
    
    return average_value, min_value

# Iterate over all Excel files in the directory
results = []

for directory in directories:
    if not os.path.exists(directory):
        # computations are still ongoing...
        continue 
    
    for filename in os.listdir(directory):
        if filename.endswith(".xlsx"):
            file_path = os.path.join(directory, filename)
            average_value, min_value = process_excel_file(file_path)
            results.append((f"{directory}-{filename}", average_value, min_value))

# Sort results in descending order of average value
results.sort(key=lambda x: x[1], reverse=True)


# Print results
for result in results[:20]:
    print(f"File: {result[0]}")
    print(f"Average: {result[1]}")
    print(f"Minimum: {result[2]}")
    print()
